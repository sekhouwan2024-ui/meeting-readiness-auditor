#!/usr/bin/env python3
"""Render a standalone meeting-readiness dashboard, simulation, XLSX book, and cheat sheet."""
from __future__ import annotations

import argparse
import html
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TYPE_LABEL = {
    "confirmed_error": "确定错误",
    "definition_risk": "口径风险",
    "logic_gap": "逻辑缺口",
    "visual_risk": "视觉风险",
    "important_omission": "重要遗漏",
    "needs_investigation": "待调查",
}
SEVERITY_LABEL = {"critical": "致命", "high": "高", "medium": "中", "low": "低"}
SEVERITY_CLASS = {"critical": "danger", "high": "danger", "medium": "warn", "low": "info"}
PRIORITY_LABEL = {"critical": "最高", "high": "高", "medium": "中", "low": "低"}
ANSWER_STATUS_LABEL = {
    "ready": "证据充分",
    "partial": "部分准备",
    "not_ready": "尚未准备",
    "decision_needed": "需要决策",
}
ANSWER_STATUS_CLASS = {
    "ready": "good",
    "partial": "warn",
    "not_ready": "danger",
    "decision_needed": "info",
}


def esc(value: Any) -> str:
    return html.escape("" if value is None else str(value))


def badge(text: str, cls: str = "neutral") -> str:
    return f'<span class="badge {cls}">{esc(text)}</span>'


def evidence_html(evidence: list[dict[str, Any]]) -> str:
    if not evidence:
        return '<div class="empty">暂无可定位证据</div>'
    rows = []
    for item in evidence:
        location = " · ".join(
            value
            for value in [item.get("file"), item.get("page_or_sheet"), item.get("cell_or_object")]
            if value
        )
        rows.append(
            '<div class="evidence-row"><div class="evidence-pin"></div><div>'
            f'<b>{esc(location)}</b><p>{esc(item.get("value_or_quote", ""))}</p></div></div>'
        )
    return "".join(rows)


def finding_card(finding: dict[str, Any]) -> str:
    severity = finding.get("severity", "medium")
    calculation = ""
    if finding.get("calculation"):
        calc = finding["calculation"]
        calculation = (
            '<div class="calc"><span>复算</span>'
            f'<code>{esc(calc.get("expression"))}</code><b>{esc(calc.get("computed_result"))}</b></div>'
        )
    return f'''
    <article id="finding-{esc(finding.get('id'))}" class="issue-card {SEVERITY_CLASS.get(severity, 'info')}" data-type="{esc(finding.get('type'))}" data-finding-id="{esc(finding.get('id'))}">
      <div class="issue-head">
        <div><span class="issue-id">{esc(finding.get('id'))}</span><h3>{esc(finding.get('title'))}</h3></div>
        <div>{badge(SEVERITY_LABEL.get(severity, severity), SEVERITY_CLASS.get(severity, 'info'))}{badge(TYPE_LABEL.get(finding.get('type'), finding.get('type')), 'neutral')}</div>
      </div>
      <div class="location">{esc(finding.get('location'))}</div>
      <div class="compare-grid">
        <div><span class="eyebrow">原表述</span><p>{esc(finding.get('statement'))}</p></div>
        <div><span class="eyebrow">核验结论</span><p>{esc(finding.get('assessment'))}</p></div>
      </div>
      {calculation}
      <details><summary>查看证据链</summary><div class="evidence-list">{evidence_html(finding.get('evidence', []))}</div></details>
      <div class="fix"><span>建议修正</span><p>{esc(finding.get('recommended_fix'))}</p></div>
    </article>'''


def question_card(question: dict[str, Any]) -> str:
    unknowns = "".join(f"<li>{esc(item)}</li>" for item in question.get("unknowns", [])) or "<li>暂无明确缺口</li>"
    return f'''
    <article id="question-{esc(question.get('id'))}" class="question-card" data-question-id="{esc(question.get('id'))}">
      <div class="question-top"><span class="qnum">{esc(question.get('id'))}</span>{badge(PRIORITY_LABEL.get(question.get('priority'), question.get('priority')), SEVERITY_CLASS.get(question.get('priority'), 'info'))}</div>
      <h3>{esc(question.get('question'))}</h3>
      <p class="why">为什么可能被问：{esc(question.get('why_it_may_be_asked'))}</p>
      <div class="answer"><span>建议直接回答</span><p>{esc(question.get('direct_answer'))}</p></div>
      <details><summary>证据与已知事实</summary><div class="evidence-list">{evidence_html(question.get('evidence', []))}</div></details>
      <details><summary>目前还不知道什么</summary><ul>{unknowns}</ul></details>
      <div class="follow"><span>下一层追问</span><p>{esc(question.get('next_question'))}</p></div>
    </article>'''


def slide_map_html(pages: list[dict[str, Any]]) -> str:
    if not pages:
        return '<div class="empty">未提供逐页风险地图</div>'
    cards = []
    for page in pages:
        count = int(page.get("issue_count", 0))
        level = page.get("risk", "clear")
        cls = {"high": "danger", "medium": "warn", "low": "info", "clear": "clear"}.get(level, "clear")
        cards.append(
            f'<div id="page-risk-P{esc(page.get("page"))}" class="page-card {cls}" data-page="{esc(page.get("page"))}"><span>P{esc(page.get("page"))}</span>'
            f'<b>{count}</b><small>{esc(page.get("title", ""))}</small></div>'
        )
    return "".join(cards)


def normalized_simulation(data: dict[str, Any]) -> dict[str, Any]:
    simulation = data.get("simulation") or {}
    if simulation.get("rounds"):
        return simulation

    roles = data.get("meeting_context", {}).get("audience_roles", ["老板"])
    rounds = []
    for index, question in enumerate(data.get("questions", [])[:8], start=1):
        unknowns = question.get("unknowns", [])
        status = "partial" if unknowns else "ready"
        rounds.append(
            {
                "id": f"S{index:02d}",
                "role": roles[(index - 1) % max(1, len(roles))],
                "question": question.get("question"),
                "intent": question.get("why_it_may_be_asked"),
                "answer_status": status,
                "suggested_answer": question.get("direct_answer"),
                "evidence": question.get("evidence", []),
                "unknowns": unknowns,
                "prep_action": "补齐未知信息后更新回答" if unknowns else "核对证据并练习口头表达",
                "follow_up_question": question.get("next_question"),
                "avoid_answer": "使用空泛表述回避核心问题",
            }
        )
    counts = Counter(round_item["answer_status"] for round_item in rounds)
    return {
        "title": "模拟过会",
        "opening": "请先用自己的话回答，再对照建议答案和证据。",
        "rounds": rounds,
        "summary": {
            "total_rounds": len(rounds),
            "ready": counts.get("ready", 0),
            "partial": counts.get("partial", 0),
            "not_ready": counts.get("not_ready", 0),
            "decision_needed": counts.get("decision_needed", 0),
        },
    }


def simulation_shell(simulation: dict[str, Any], standalone: bool = False) -> str:
    summary = simulation.get("summary", {})
    rounds_json = json.dumps(simulation.get("rounds", []), ensure_ascii=False).replace("</", "<\\/")
    shell_class = "simulation-shell standalone-shell" if standalone else "simulation-shell"
    return f'''
    <div id="simulation-console" class="{shell_class}" data-simulation-root>
      <div class="sim-head">
        <div>
          <span class="eyebrow">LIVE REHEARSAL</span>
          <h2>{esc(simulation.get('title', '模拟过会'))}</h2>
          <p>{esc(simulation.get('opening', ''))}</p>
        </div>
        <div class="sim-summary">
          <span><b>{esc(summary.get('total_rounds', 0))}</b>轮追问</span>
          <span class="good"><b>{esc(summary.get('ready', 0))}</b>证据充分</span>
          <span class="warn"><b>{esc(summary.get('partial', 0))}</b>部分准备</span>
          <span class="danger"><b>{esc(summary.get('not_ready', 0))}</b>尚未准备</span>
          <span class="info"><b>{esc(summary.get('decision_needed', 0))}</b>需要决策</span>
        </div>
      </div>
      <div id="simulation-stage" class="sim-stage" data-sim-stage>
        <div class="sim-stage-top">
          <div class="role-chip" data-sim-role>提问角色</div>
          <div class="sim-progress-label"><span data-sim-index>1</span> / <span data-sim-total>0</span></div>
        </div>
        <div class="sim-progress"><i data-sim-progress></i></div>
        <div class="sim-intent" data-sim-intent></div>
        <h3 id="simulation-question" class="sim-question" data-sim-question>正在载入问题……</h3>
        <label class="answer-box">
          <span>先用自己的话回答</span>
          <textarea data-sim-user-answer placeholder="不要先看标准答案。用20—45秒能说完的方式回答。"></textarea>
        </label>
        <div class="sim-actions">
          <button class="primary" type="button" data-sim-reveal>对照建议回答</button>
          <button class="secondary" type="button" data-sim-next disabled>下一问</button>
          <button class="ghost" type="button" data-sim-restart>重新开始</button>
        </div>
        <div id="simulation-feedback" class="sim-feedback" data-sim-feedback hidden>
          <div class="feedback-grid">
            <div id="simulation-answer" class="feedback-main" data-sim-answer-region>
              <div class="feedback-title"><span data-sim-status class="status-pill"></span><b>建议直接回答</b></div>
              <p data-sim-answer></p>
            </div>
            <div id="simulation-action" class="feedback-side" data-sim-action-region>
              <span>会前动作</span><p data-sim-action></p>
            </div>
          </div>
          <div class="feedback-columns">
            <div id="simulation-evidence"><span class="eyebrow">证据</span><div data-sim-evidence></div></div>
            <div id="simulation-unknowns"><span class="eyebrow">当前未知</span><ul data-sim-unknowns></ul></div>
          </div>
          <div id="simulation-avoid" class="avoid-answer"><span>避免这样答</span><p data-sim-avoid></p></div>
          <div id="simulation-followup" class="next-pressure"><span>下一层追问</span><p data-sim-followup></p></div>
        </div>
      </div>
      <script type="application/json" data-sim-data>{rounds_json}</script>
    </div>'''


def render_html(data: dict[str, Any], out: Path, css: str, js: str):
    findings = data.get("findings", [])
    questions = data.get("questions", [])
    requests = data.get("data_requests", [])
    simulation = normalized_simulation(data)
    counts = Counter(finding.get("type") for finding in findings)
    meeting_status = data.get("summary", {}).get("meeting_status")
    if not meeting_status:
        if counts.get("confirmed_error", 0):
            meeting_status = "needs_fix"
        elif requests:
            meeting_status = "needs_data"
        else:
            meeting_status = "ready_for_rehearsal"
    status_label = {
        "needs_fix": "先完成会前处理",
        "needs_data": "会前仍需补数",
        "ready_for_rehearsal": "可以进入答辩演练",
        "ready_with_caveats": "可演练，但保留边界",
    }.get(meeting_status, "需要人工复核")
    status_class = {
        "needs_fix": "danger",
        "needs_data": "warn",
        "ready_for_rehearsal": "good",
        "ready_with_caveats": "info",
    }.get(meeting_status, "info")
    issue_cards = "".join(finding_card(finding) for finding in findings) or '<div class="empty">未发现明确修正项</div>'
    question_cards = "".join(question_card(question) for question in questions) or '<div class="empty">未生成高风险追问</div>'
    request_rows = "".join(
        f'<tr id="gap-D{index:02d}" data-gap-id="D{index:02d}"><td>{esc(item.get("item"))}</td><td>{esc(item.get("reason"))}</td><td>{esc(item.get("priority"))}</td><td>{esc(item.get("owner"))}</td><td>{esc(item.get("deadline"))}</td></tr>'
        for index, item in enumerate(requests, start=1)
    ) or '<tr><td colspan="5">暂无会前补数项</td></tr>'
    top_questions = "".join(
        f'<div id="cheat-{esc(question.get("id"))}" class="cheat-row" data-question-id="{esc(question.get("id"))}"><b>{esc(question.get("question"))}</b><p>{esc(question.get("direct_answer"))}</p></div>'
        for question in questions[:6]
    ) or '<div class="empty">暂无答辩条目</div>'

    body = f'''<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{esc(data.get('report_title'))}</title><style>{css}</style></head>
<body>
<header class="hero"><div class="brand"><span class="brand-mark">会</span><div><b>过会排雷</b><small>Meeting Readiness Console</small></div></div>
<div class="hero-copy"><span class="eyebrow">{esc(data.get('meeting_context', {}).get('meeting_type', '业务汇报'))}</span><h1>{esc(data.get('report_title'))}</h1><p>{esc(data.get('summary', {}).get('executive_summary', ''))}</p></div>
<div class="hero-status {status_class}"><span>当前会前状态</span><b>{esc(status_label)}</b><small>{counts.get('confirmed_error',0)}项确定修正 · {len(requests)}项待补资料</small></div></header>
<nav class="tabs"><button class="active" data-tab="overview">总览</button><button data-tab="simulation">模拟过会</button><button data-tab="fixes">会前处理</button><button data-tab="questions">追问清单</button><button data-tab="cheat">答辩小抄</button><button data-tab="data">会前补数</button></nav>
<main>
<section id="overview" class="tab-panel active">
<div class="kpi-grid">
<div class="kpi danger"><span>确定错误</span><b>{counts.get('confirmed_error',0)}</b></div>
<div class="kpi warn"><span>口径与逻辑风险</span><b>{counts.get('definition_risk',0)+counts.get('logic_gap',0)}</b></div>
<div class="kpi info"><span>披露与待调查</span><b>{counts.get('visual_risk',0)+counts.get('important_omission',0)+counts.get('needs_investigation',0)}</b></div>
<div class="kpi neutral"><span>模拟追问</span><b>{simulation.get('summary',{}).get('total_rounds',0)}</b></div>
</div>
<div class="panel"><div class="panel-head"><div><span class="eyebrow">PAGE RISK MAP</span><h2>逐页雷区地图</h2></div></div><div class="page-map">{slide_map_html(data.get('page_risk_map', []))}</div></div>
<div class="two-col"><div class="panel"><span class="eyebrow">处理优先级</span><h2>先处理什么</h2><ol class="priority-list">{''.join(f'<li><b>{esc(f.get("title"))}</b><span>{esc(f.get("location"))}</span></li>' for f in findings[:5]) or '<li>未发现需要立即纠正的明显问题</li>'}</ol></div>
<div class="panel"><span class="eyebrow">会议重点</span><h2>最可能被追问</h2><ol class="priority-list">{''.join(f'<li><b>{esc(q.get("question"))}</b><span>{esc(q.get("why_it_may_be_asked"))}</span></li>' for q in questions[:5]) or '<li>暂无追问</li>'}</ol></div></div>
</section>
<section id="simulation" class="tab-panel"><div class="section-title"><span class="eyebrow">REHEARSE BEFORE MEETING</span><h2>把明天的汇报，提前开一遍</h2><p>先自行回答，再对照证据、未知信息和下一层追问。系统不对自由输入制造虚假评分。</p></div>{simulation_shell(simulation)}</section>
<section id="fixes" class="tab-panel"><div class="section-title"><span class="eyebrow">FIX BEFORE MEETING</span><h2>先处理可确认问题，再进入答辩推演</h2><p>明确错误应先修正；可比性、分群、归因和证据边界应在会议前说明。</p></div><div class="issue-list">{issue_cards}</div></section>
<section id="questions" class="tab-panel"><div class="section-title"><span class="eyebrow">CHALLENGE LIST</span><h2>基于修正后事实的高风险追问</h2><p>问题聚焦原因、判断、取舍、行动和风险，而不是重复追问已修正的数据错误。</p></div><div class="question-grid">{question_cards}</div></section>
<section id="cheat" class="tab-panel"><div class="section-title"><span class="eyebrow">ONE-PAGE BRIEF</span><h2>会前答辩小抄</h2><p>先直接回答，再给证据；不知道的部分明确说明并给补数计划。</p></div><div class="cheat-sheet">{top_questions}</div></section>
<section id="data" class="tab-panel"><div class="section-title"><span class="eyebrow">DATA GAPS</span><h2>会前补数清单</h2><p>数据正确不代表数据充分。以下信息会影响关键结论或回答完整性。</p></div><div class="panel table-wrap"><table><thead><tr><th>需要补充</th><th>为什么需要</th><th>优先级</th><th>负责人</th><th>截止时间</th></tr></thead><tbody>{request_rows}</tbody></table></div></section>
</main>
<footer>生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')} · AI结果必须由汇报负责人复核，不得据此编造数据或掩盖风险。</footer>
<script>{js}</script></body></html>'''
    out.write_text(body, encoding="utf-8")


def render_simulation_html(data: dict[str, Any], out: Path, css: str, js: str):
    simulation = normalized_simulation(data)
    body = f'''<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
<title>{esc(simulation.get('title', '模拟过会'))}</title><style>{css}</style></head>
<body class="simulation-page">
<header class="simulation-brand"><div class="brand"><span class="brand-mark">会</span><div><b>模拟过会</b><small>由「过会排雷」Skill生成</small></div></div><p title="{esc(data.get('report_title'))}">{esc(data.get('report_title'))}</p></header>
<main class="simulation-main">{simulation_shell(simulation, standalone=True)}</main>
<footer>先核验事实与证据边界，再练习回答。AI建议必须由汇报负责人核验。</footer>
<script>{js}</script></body></html>'''
    out.write_text(body, encoding="utf-8")


def style_sheet(ws, widths: list[int]):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    dark = "172033"
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    thin = Side(style="thin", color="D9E0EA")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)


def render_xlsx(data: dict[str, Any], out: Path):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "会前处理"
    ws.append(["编号", "页码/位置", "类型", "风险", "原表述", "核验结论", "建议修正", "证据"])
    for finding in data.get("findings", []):
        evidence = "\n".join(
            " / ".join(
                value
                for value in [item.get("file"), item.get("page_or_sheet"), item.get("cell_or_object"), str(item.get("value_or_quote", ""))]
                if value
            )
            for item in finding.get("evidence", [])
        )
        ws.append(
            [
                finding.get("id"),
                finding.get("location"),
                TYPE_LABEL.get(finding.get("type"), finding.get("type")),
                SEVERITY_LABEL.get(finding.get("severity"), finding.get("severity")),
                finding.get("statement"),
                finding.get("assessment"),
                finding.get("recommended_fix"),
                evidence,
            ]
        )
    style_sheet(ws, [10, 20, 14, 10, 32, 38, 38, 48])

    qws = workbook.create_sheet("高风险追问")
    qws.append(["编号", "优先级", "可能追问", "为什么会问", "建议直接回答", "证据", "目前未知", "下一层追问"])
    for question in data.get("questions", []):
        evidence = "\n".join(
            " / ".join(
                value
                for value in [item.get("file"), item.get("page_or_sheet"), item.get("cell_or_object"), str(item.get("value_or_quote", ""))]
                if value
            )
            for item in question.get("evidence", [])
        )
        qws.append(
            [
                question.get("id"),
                PRIORITY_LABEL.get(question.get("priority"), question.get("priority")),
                question.get("question"),
                question.get("why_it_may_be_asked"),
                question.get("direct_answer"),
                evidence,
                "\n".join(question.get("unknowns", [])),
                question.get("next_question"),
            ]
        )
    style_sheet(qws, [10, 10, 36, 34, 48, 48, 34, 36])

    sws = workbook.create_sheet("模拟过会")
    sws.append(["轮次", "角色", "问题", "意图", "准备状态", "建议回答", "证据", "当前未知", "会前动作", "下一层追问", "避免这样答"])
    for round_item in normalized_simulation(data).get("rounds", []):
        evidence = "\n".join(
            " / ".join(
                value
                for value in [item.get("file"), item.get("page_or_sheet"), item.get("cell_or_object"), str(item.get("value_or_quote", ""))]
                if value
            )
            for item in round_item.get("evidence", [])
        )
        sws.append(
            [
                round_item.get("id"),
                round_item.get("role"),
                round_item.get("question"),
                round_item.get("intent"),
                ANSWER_STATUS_LABEL.get(round_item.get("answer_status"), round_item.get("answer_status")),
                round_item.get("suggested_answer"),
                evidence,
                "\n".join(round_item.get("unknowns", [])),
                round_item.get("prep_action"),
                round_item.get("follow_up_question"),
                round_item.get("avoid_answer"),
            ]
        )
    style_sheet(sws, [10, 12, 38, 34, 14, 52, 48, 34, 38, 38, 32])

    dws = workbook.create_sheet("会前补数")
    dws.append(["需要补充", "原因", "优先级", "负责人", "截止时间"])
    for request in data.get("data_requests", []):
        dws.append([request.get("item"), request.get("reason"), request.get("priority"), request.get("owner"), request.get("deadline")])
    style_sheet(dws, [32, 48, 12, 20, 18])
    workbook.save(out)


def render_markdown(data: dict[str, Any], out: Path):
    lines = [
        f"# {data.get('report_title', '过会答辩小抄')}",
        "",
        f"> {data.get('summary', {}).get('executive_summary', '')}",
        "",
        "## 模拟过会",
    ]
    for round_item in normalized_simulation(data).get("rounds", [])[:8]:
        lines.extend(
            [
                "",
                f"### {round_item.get('role')}：{round_item.get('question')}",
                f"**准备状态：** {ANSWER_STATUS_LABEL.get(round_item.get('answer_status'), round_item.get('answer_status'))}",
                f"**建议回答：** {round_item.get('suggested_answer')}",
                f"**会前动作：** {round_item.get('prep_action')}",
            ]
        )
        if round_item.get("unknowns"):
            lines.append("**尚未确认：** " + "；".join(round_item.get("unknowns")))
        lines.append(f"**下一层追问：** {round_item.get('follow_up_question')}")
    lines.extend(["", "## 会前必须补充"])
    for request in data.get("data_requests", []):
        lines.append(f"- **{request.get('item')}**：{request.get('reason')}（{request.get('priority')}，{request.get('deadline')}）")
    out.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("audit_json")
    parser.add_argument("--out", required=True)
    args = parser.parse_args()
    skill_dir = Path(__file__).resolve().parents[1]
    data = json.loads(Path(args.audit_json).read_text(encoding="utf-8"))
    output_dir = Path(args.out).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    css = (skill_dir / "assets" / "report-style.css").read_text(encoding="utf-8")
    js = (skill_dir / "assets" / "report-script.js").read_text(encoding="utf-8")

    report_path = output_dir / "过会排雷报告.html"
    simulation_path = output_dir / "模拟过会.html"
    workbook_path = output_dir / "会前处理与答辩准备.xlsx"
    cheat_path = output_dir / "一页答辩小抄.md"
    render_html(data, report_path, css, js)
    render_simulation_html(data, simulation_path, css, js)
    render_xlsx(data, workbook_path)
    render_markdown(data, cheat_path)
    print(
        json.dumps(
            {
                "html": str(report_path),
                "simulation": str(simulation_path),
                "xlsx": str(workbook_path),
                "cheatsheet": str(cheat_path),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
